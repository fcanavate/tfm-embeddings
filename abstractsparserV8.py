import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import nltk, re, pprint
nltk.download('punkt')
from nltk import word_tokenize	
from nltk.tokenize import sent_tokenize
from urllib import request
import parserESV3
import numpy as np
from scipy import sparse 
from scipy import stats
nltk.download('averaged_perceptron_tagger')
origen=""
row=0


wb=load_workbook(filename="verbnoun.xlsx")
ws=wb.active
n=0
wb_name=load_workbook(filename="base_noun.xlsx")
ws_name=wb_name.active
wb_verb=load_workbook(filename="base_verb.xlsx")
ws_verb=wb_verb.active
wb_coord_nounsuj_verb=load_workbook(filename="coord_nounsuj_verb.xlsx")
ws_coord_nounsuj_verb=wb_coord_nounsuj_verb.active
wb_coord_nounobj_verb=load_workbook(filename="coord_nounobj_verb.xlsx")
ws_coord_nounobj_verb=wb_coord_nounobj_verb.active
wb_coord_nounsuj_obj_verb=load_workbook(filename="coord_nounsuj_obj_verb.xlsx")
ws_coord_nounsuj_obj_verb=wb_coord_nounsuj_obj_verb.active
wb_coord_nounsuj_obj_verb_text=load_workbook(filename="coord_nounsuj_obj_verb_text.xlsx")
ws_coord_nounsuj_obj_verb_text=wb_coord_nounsuj_obj_verb_text.active

## funcion base_verb:
# argumento: verbo 
# añade verbo a verbES, devuelve fila, sea la última si se ha añadido o existente si existía
def base_verb(verb,forma):
    last_row=ws_verb.max_row
    if ws_verb.cell(last_row,1).value:
        last_id=ws_verb.cell(last_row,1).value
    else:
        last_id=0
    if last_id=="":
        last_id=0
    #print ("baseverb_last_row",last_row)
    resultado=True
    row=1
    for row in range(1,last_row+1):
        celda_id=ws_verb.cell(row,1).value
        if ws_verb.cell(row,1).value:
            celda_id=ws_verb.cell(row,1).value
        else:
            celda_id=0
        if celda_id=="":
            celda_id=0
        celda=ws_verb.cell(row,2).value
        if celda==verb:
            resultado=False
            return(celda_id)
    if resultado:
        #print("verboañadido:",verb)
        ws_verb.append([last_id+1,verb,forma])
        return(last_id+1)

## funcion base_noun:
# argumento: nombre 
# añade nombre a base_noun, devuelve fila, sea la última si se ha aadido o exitente si existía
def base_noun(nombre,forma):
    last_row=ws_name.max_row
    last_id=ws_name.cell(last_row,1).value
    if ws_name.cell(last_row,1).value:
        last_id=ws_name.cell(last_row,1).value
    else:
        last_id=0
    if last_id=="":
        last_id=0
    #print ("basenoun_last_row",last_row+1)
    resultado=True
    row=0
    for row in range(1,last_row+1):
        celda_id=ws_name.cell(row,1).value
        if ws_name.cell(row,1).value:
            celda_id=ws_name.cell(row,1).value
        else:
            celda_id=0
        if celda_id=="":
            celda_id=0
        celda=ws_name.cell(row,2).value
        if celda==nombre:
            resultado=False
            return(celda_id)
    if resultado:
        ws_name.append([last_id+1,nombre,forma])    
        return(last_id+1)

## funcion coord_nounsuj_verb
# ARGUMENTOS: (indexverb=numero de columna correspondiente verb de baseverb)
# devuelve: la fila número indexverb que tiene por elementos los que tenia la fila en la matriz contadorverbsuj +1 en el elemento correspondiente a indexnounsuj            
def coord_nounsuj_verb(indexnounsuj,indexverb):
    coord_nounsuj_verb_fila=[indexnounsuj,indexverb,1]
    ws_coord_nounsuj_verb.append(coord_nounsuj_verb_fila)
## funcion coord_nounobj_verb
# ARGUMENTOS: (indexverb=numero de columna correspondiente verb de baseverb)
# devuelve: la fila número indexverb que tiene por elementos los que tenia la fila en la matriz contadorverbsuj +1 en el elemento correspondiente a indexnounsuj            
def coord_nounobj_verb(indexnounobj,indexverb):
    coord_nounobj_verb_fila=[indexnounobj,indexverb,1]
    ws_coord_nounobj_verb.append(coord_nounobj_verb_fila)
def coord_nounsuj_obj_verb(index_noun_suj,index_noun_obj,indexverb):
    coord_nounsuj_obj_verb_fila=[index_noun_suj,index_noun_obj,indexverb,origen,row]
    ws_coord_nounsuj_obj_verb.append(coord_nounsuj_obj_verb_fila)
def coord_nounsuj_obj_verb_text(index_noun_suj,index_noun_obj,indexverb,noun_suj,noun_obj,verb,oracion):
    coord_nounsuj_obj_verb_fila_text=[index_noun_suj,index_noun_obj,indexverb,origen,row,noun_suj,noun_obj,verb,oracion]
    ws_coord_nounsuj_obj_verb_text.append(coord_nounsuj_obj_verb_fila_text)
def parsergeneral(texto):
    sentences=texto
    nsentence=0
    #for sentence in sentences[1:5000]:
    for sentence in sentences:
        nsentence+=1
        #print ("sentencia nº=",nsentence)
        #print (sentence)
        sentence1=sentence.strip("\r\n")
        sentence2=sentence1.replace("\r\n"," ")
        sentence3=sentence2.replace("  "," ")
        sentence3=sentence3.translate ({ord(c): " " for c in "@#$%^&*()[]{};:/<>\|`~-=_+"})
        #llamada al parse
        #print(sentence3)
        analisis=parserESV3.parse(sentence3)
        #print ("analisis",analisis)
        npalabra=1
        for palabra in analisis:
            if len(palabra)>6:
                if palabra[3]=='VERB':
                    #print("verbo",palabra[2])
                    verbnoun=["","","","","","","","","","",""]
                    verbnoun[0]=palabra[2]
                    verbnoun[10]=sentence3
                    ordenverbo=palabra[0]
                    npalabra1=0
                    npalabrasuj=1
                    npalabraobj=5
                    sujetodetectado=False
                    index_verb=base_verb(palabra[2],palabra[1])
                    verbo_raiz=palabra[2]
                    for palabra1 in analisis:
                        #print ("palabra1:",palabra1)
                        npalabra1=npalabra1+1
                        if len(palabra1)>6:             
                            if npalabrasuj<11 and palabra1[6]==ordenverbo and (palabra1[3]=="NOUN" or palabra1[3]=="PROPN") and palabra1[7]=="nsubj":
                                verbnoun[npalabrasuj]=palabra1[2]
                                index_noun=base_noun(palabra1[2],palabra1[1])
                                coord_nounsuj_verb(index_noun,index_verb)
                                npalabrasuj=npalabrasuj+1
                                sujetodetectado=True
                                nombre_suj=palabra1[2]
                                index_noun_suj=index_noun
                            if npalabraobj<11 and palabra1[6]==ordenverbo and (palabra1[3]=="NOUN" or palabra1[3]=="PROPN") and palabra1[7]=="obj":
                                verbnoun[npalabraobj]=palabra1[2]
                                index_noun=base_noun(palabra1[2],palabra1[1])
                                coord_nounobj_verb(index_noun,index_verb)
                                npalabraobj=npalabraobj+1
                                if sujetodetectado:
                                    coord_nounsuj_obj_verb(index_noun_suj,index_noun,index_verb)
                                    coord_nounsuj_obj_verb_text(index_noun_suj,index_noun,index_verb,nombre_suj,palabra1[2],verbo_raiz,sentence3)

                    #print ("verbnoun=",verbnoun)
                    if npalabrasuj>1 or npalabraobj>5:
                        ws.append(verbnoun)

#
# recorre los resumenes y los parsea
numerosent=0
for libroindex in range(1,11): 
    print ("procesando libro",libroindex)                       
    libro=load_workbook(filename="wikipedia0_"+str(libroindex)+"00000.xlsx")
    hoja=libro.active
    maxrow=hoja.max_row
    for row in range(1,maxrow):
        origen=hoja.cell(row=row,column=1).value
        resumen=hoja.cell(row=row,column=3)
        texto = resumen.value
        #print (texto)
        print ("numero resumen:",row," del numero de libro:", libroindex)
        if texto != None:
            sentences=sent_tokenize(texto)
            numerosent+=len(sentences)
            print ("numerosentencias:",numerosent)
            parsergeneral(sentences)
        if row==5000:
                wb_coord_nounsuj_verb.save("coord_nounsuj_verb_prueba.xlsx")
                wb_coord_nounobj_verb.save("coord_nounobj_verb_prueba.xlsx")
                wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb_prueba.xlsx")
                wb_coord_nounsuj_obj_verb_text.save("coord_nounsuj_obj_verb_prueba_text.xlsx")
                wb.save("verbnoun_prueba.xlsx")
                wb=load_workbook(filename="verbnoun_prueba.xlsx")
                ws=wb.active
                wb_verb.save("base_verb_prueba.xlsx")
                wb_name.save("base_noun_prueba.xlsx")
        if row==50000:
                wb_coord_nounsuj_verb.save("coord_nounsuj_verb"+str(libroindex*100000-50000)+".xlsx")
                wb_coord_nounobj_verb.save("coord_nounobj_verb"+str(libroindex*100000-50000)+".xlsx")
                wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb"+str(libroindex*100000-50000)+".xlsx")
                wb_coord_nounsuj_obj_verb_text.save("coord_nounsuj_obj_verb_text"+str(libroindex*100000-50000)+".xlsx")
                wb.save("verbnoun"+str(libroindex*100000-50000)+".xlsx")
                wb=load_workbook(filename="verbnoun.xlsx")
                ws=wb.active
                wb_verb.save("base_verb"+str(libroindex*100000-50000)+".xlsx")
                wb_name.save("base_noun"+str(libroindex*100000-50000)+".xlsx")
    wb_verb.save("base_verb.xlsx")
    wb_name.save("base_noun.xlsx")
    wb_coord_nounsuj_verb.save("coord_nounsuj_verb"+str(libroindex)+"00000.xlsx")
    wb_coord_nounobj_verb.save("coord_nounobj_verb"+str(libroindex)+"00000.xlsx")
    wb_coord_nounobj_verb.save("coord_nounsuj_obj"+str(libroindex)+"00000.xlsx")
    wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb"+str(libroindex)+"00000.xlsx")
    wb_coord_nounsuj_obj_verb_text.save("coord_nounsuj_obj_verb_text"+str(libroindex)+"00000.xlsx")
    wb.save("verbnoun"+str(libroindex)+"00000.xlsx")
    wb=load_workbook(filename="verbnoun.xlsx")
    ws=wb.active
    wb_coord_nounsuj_verb=load_workbook(filename="coord_nounsuj_verb.xlsx")
    ws_coord_nounsuj_verb=wb_coord_nounsuj_verb.active
    wb_coord_nounobj_verb=load_workbook(filename="coord_nounobj_verb.xlsx")
    ws_coord_nounobj_verb=wb_coord_nounobj_verb.active
    wb_verb.save("base_verb"+str(libroindex)+"00000.xlsx")
    wb_name.save("base_noun"+str(libroindex)+"00000.xlsx")
    print ("numero sentencias total grabadas:",numerosent)
    print ("libro grabado:",libroindex) 
wb_verb.save("base_verb.xlsx")
wb_name.save("base_noun.xlsx")
wb_coord_nounsuj_verb.save("coord_nounsuj_verb"+str(libroindex)+"00000.xlsx")
wb_coord_nounobj_verb.save("coord_nounobj_verb"+str(libroindex)+"00000.xlsx")
wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb"+str(libroindex)+"00000.xlsx")
wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb_text"+str(libroindex)+"00000.xlsx")
wb.save("verbnoun"+str(libroindex)+"00000.xlsx")
wb=load_workbook(filename="verbnoun.xlsx")
ws=wb.active
wb_coord_nounsuj_verb=load_workbook(filename="coord_nounsuj_verb.xlsx")
ws_coord_nounsuj_verb=wb_coord_nounsuj_verb.active
wb_coord_nounobj_verb=load_workbook(filename="coord_nounobj_verb.xlsx")
ws_coord_nounobj_verb=wb_coord_nounobj_verb.active
wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb"+str(libroindex)+"00000.xlsx")
wb_coord_nounsuj_obj_verb.save("coord_nounsuj_obj_verb_text"+str(libroindex)+"00000.xlsx")
wb_verb.save("base_verb"+str(libroindex)+"00000.xlsx")
wb_name.save("base_noun"+str(libroindex)+"00000.xlsx")
wb_coord_nounobj_verb.save("coord_nounsuj_obj"+str(libroindex)+"00000.xlsx")
print ("numero sentencias total grabadas:",numerosent)
print ("ultimo ibro grabado:",libroindex) 
                    
#completar()
# coo_matrix.sum_duplicates()[source]   # esto elimina celdas duplicadas y suma sus valores
# creo que no hace falta lo anterior. siplemnte al definirla con indices duplicados los suma. ver otros metodos ttps://docs.scipy.org/doc/scipy/reference/generated/scipy.sparse.coo_matrix.html
# falta implementar la obtnecion al principio y cada 1000 frases la grabacion a excel.